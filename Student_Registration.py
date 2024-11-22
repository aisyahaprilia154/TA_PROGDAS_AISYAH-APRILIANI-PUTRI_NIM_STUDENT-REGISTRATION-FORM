from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
import re  # Import untuk regex
from tkinter.ttk import Treeview

undo_stack = []

def save_state():
    """Menyimpan state form ke dalam stack."""
    current_state = {
        "Name": Name.get(),
        "TL": TL.get(),
        "Tempat": Tempat.get(),
        "Alamat": Alamat.get(),
        "Email": Email.get(),
        "NoHP": NoHP.get(),
        "F_Name": F_Name.get(),
        "M_Name": M_Name.get(),
        "Father_Occupation": Father_Occupation.get(),
        "Mother_Occupation": Mother_Occupation.get(),
        "Jurusan1": Jurusan1.get(),
        "Jurusan2": Jurusan2.get(),
        "Gender": radio.get(),
    }
    undo_stack.append(current_state)

def load_state(state):
    """Memuat state form dari stack."""
    Name.set(state["Name"])
    TL.set(state["TL"])
    Tempat.set(state["Tempat"])
    Alamat.set(state["Alamat"])
    Email.set(state["Email"])
    NoHP.set(state["NoHP"])
    F_Name.set(state["F_Name"])
    M_Name.set(state["M_Name"])
    Father_Occupation.set(state["Father_Occupation"])
    Mother_Occupation.set(state["Mother_Occupation"])
    Jurusan1.set(state["Jurusan1"])
    Jurusan2.set(state["Jurusan2"])
    radio.set(state["Gender"])

def undo():
    """Membatalkan perubahan terakhir pada form."""
    if undo_stack:
        last_state = undo_stack.pop()  # Ambil state terakhir
        load_state(last_state)
    else:
        messagebox.showinfo("Undo", "Tidak ada perubahan untuk dibatalkan!")

if not os.path.exists("Student Images"):
    os.makedirs("Student Images")

background = "#2E4053"
framebg = "#D5D8DC"
framefg = "#2C3E50"
font_style = ("Verdana", 10)

root = Tk()
root.title("Student Registration System")
root.geometry("1300x750+100+50")
root.config(bg=background)

# File excel
file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active  # Menambahkan inisialisasi sheet
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Nama"
    sheet['C1'] = "Tanggal lahir"
    sheet['D1'] = "Tempat lahir"
    sheet['E1'] = "Alamat"
    sheet['F1'] = "Date Of Registration"
    sheet['G1'] = "Gender"
    sheet['H1'] = "No HP"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    sheet['M1'] = "Email"

    file.save('Student_data.xlsx')

# Button 
def Exit():
    root.destroy()

def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(
        initialdir=os.getcwd(),
        title="Pilih file gambar",
        filetypes=[("JPG File", "*.jpg"), ("PNG File", "*.png"), 
                   ("JPEG File", "*.jpeg"), ("BMP File", "*.bmp"), 
                   ("GIF File", "*.gif"), ("All files", "*.*")]
    )

    if filename:
        img = Image.open(filename)
        resized_image = img.resize((190, 190))
        photo2 = ImageTk.PhotoImage(resized_image)
        
        lbl.config(image=photo2)
        lbl.image = photo2

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    try:
        max_row_value = int(sheet.cell(row=row, column=1).value)
        Registration.set(str(max_row_value + 1))
    except:
        Registration.set("1")

def Clear():
    save_state()
    global img
    Name.set('')
    TL.set('')
    Tempat.set('')
    Alamat.set('')
    Email.set('')
    NoHP.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Jurusan1.set("Pilihan 1")
    Jurusan2.set("Pilihan 2")
    
    radio.set(0)

    saveButton.config(state='normal')

    try:
        img1 = Image.open('Images/upload photo.png')
        img1_resized = img1.resize((190, 190), Image.LANCZOS)
        img1_resized = ImageTk.PhotoImage(img1_resized)
        lbl.config(image=img1_resized)
        lbl.image = img1_resized  
        img = img1_resized  
    except Exception as e:
        print(f"Error saat mengganti gambar: {e}")

    registration_no()

#validasi nama
def is_valid_name(name):
    return bool(re.match("^[A-Za-z ]+$", name))
#validasi nomor
def is_valid_number(number):
    return bool(re.match("^[0-9]+$", number))

def Save():
    R1 = Registration.get()
    N1 = Name.get()
    T1 = TL.get()
    T2 = Tempat.get()
    A1 = Alamat.get()
    E1 = Email.get()
    N2 = NoHP.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = Father_Occupation.get()
    M1 = Mother_Occupation.get()
    J1 = Jurusan1.get()
    J2 = Jurusan2.get()
    D1 = Date.get()

    # Validasi apakah semua field sudah diisi
    if not all([R1, N1, T1, T2, A1, E1, N2, fathername, mothername, F1, M1, J1, J2]):
        messagebox.showerror("Error", "Semua field harus diisi!")
        return

    # Validasi nomor registrasi (hanya angka)
    if not is_valid_number(R1):
        messagebox.showerror("Error", "Nomor registrasi hanya boleh diisi angka!")
        return

    # Validasi nama (hanya huruf dan spasi)
    if not is_valid_name(N1):
        messagebox.showerror("Error", "Nama hanya boleh diisi huruf!")
        return

    # Validasi No HP (hanya angka)
    if not is_valid_number(N2):
        messagebox.showerror("Error", "No HP hanya boleh diisi angka!")
        return

    # Validasi nama ayah (hanya huruf dan spasi)
    if not is_valid_name(fathername):
        messagebox.showerror("Error", "Nama Ayah hanya boleh diisi huruf!")
        return

    # Validasi nama ibu (hanya huruf dan spasi)
    if not is_valid_name(mothername):
        messagebox.showerror("Error", "Nama Ibu hanya boleh diisi huruf!")
        return

    # Validasi pekerjaan ayah (hanya huruf dan spasi)
    if not is_valid_name(F1):
        messagebox.showerror("Error", "Pekerjaan Ayah hanya boleh diisi huruf!")
        return

    # Validasi pekerjaan ibu (hanya huruf dan spasi)
    if not is_valid_name(M1):
        messagebox.showerror("Error", "Pekerjaan Ibu hanya boleh diisi huruf!")
        return

    # Validasi gender
    G1 = ""
    if radio.get() == 1:
        G1 = "Male"
    elif radio.get() == 2:
        G1 = "Female"
    else:
        messagebox.showerror("Error", "Pilih gender!")
        return

    # Validasi gambar
    if 'img' not in globals():
        messagebox.showerror("Error", "Gambar wajib di-upload!")
        return

    try:
        # Load file Excel
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active

        # Membuat row baru untuk diisi dengan data
        new_row = [R1, N1, T1, T2, A1, D1, G1, N2, fathername, mothername, F1, M1, E1]

        # Menyimpan data ke Excel
        sheet.append(new_row)
        file.save('Student_data.xlsx')
        file.close()

        try:
            # Menyimpan gambar ke folder 
            img.save(f"Student Images/{R1}.jpg")
            messagebox.showinfo("Info", "Data dan gambar berhasil disimpan!")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan gambar: {e}")

        # Membersihkan form setelah data disimpan
        Clear()
        registration_no()
    except Exception as e:
        messagebox.showerror("Error", f"Gagal menyimpan data: {e}")

#gender
def selection():
    value=radio.get()
    if value==1:
        gender="Male"
        print(gender)
    else:
        gender="Female"
        print(gender)
        
#top frames
Label (root,text="Email: aisyahaprilia1515@gmail.com",width=10,height=3,bg="#FFFFFF",anchor='e').pack(side= TOP,fill=X)
Label (root,text="STUDENT REGISTRATION",width=10,height=2,bg="#0078C8",fg='#fff',font='arial 20 bold' ).pack(side= TOP,fill=X)

# Registration and Date
Label(root, text="No Registrasi:", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date:", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

Registration = StringVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)

Date.set(d1)

# Student details
obj=LabelFrame(root,text="Data Pribadi",font=20,bd=2,width=1000,bg=framebg,fg=framefg,height=350,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Nama Lengkap:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Tanggal lahir:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Tempat lahir:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)
Label(obj,text="Alamat:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=200)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=250)


Label(obj,text="Email:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="No HP:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Jurusan1:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)
Label(obj,text="Jurusan2:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=200)

Name=StringVar()
name_entry = Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

TL=StringVar()
tl_entry = Entry(obj,textvariable=TL,width=20,font="arial 10")
tl_entry.place(x=160,y=100)

Tempat=StringVar()
tempat_entry = Entry(obj,textvariable=Tempat,width=20,font="arial 10")
tempat_entry.place(x=160,y=150)

Alamat=StringVar()
alamat_entry = Entry(obj,textvariable=Alamat,width=20,font="arial 10")
alamat_entry.place(x=160,y=200)

radio= IntVar()
R1 = Radiobutton(obj,text="Male", variable=radio, value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=250,y=250)

R2 = Radiobutton(obj,text="Female", variable=radio, value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=160,y=250)

Email=StringVar()
email_entry = Entry(obj,textvariable=Email,width=20,font="arial 10")
email_entry.place(x=600,y=50)

NoHP=StringVar()
nohp_entry = Entry(obj,textvariable=NoHP,width=20,font="arial 10")
nohp_entry.place(x=600,y=100)


# Daftar jurusan di Undip
jurusan_undip = [
    "Fakultas Teknik - Teknik Sipil", "Fakultas Teknik - Arsitektur", 
    "Fakultas Teknik - Teknik Mesin", "Fakultas Teknik - Teknik Kimia", 
    "Fakultas Teknik - Teknik Elektro", "Fakultas Teknik - Teknik Industri",
    "Fakultas Teknik - Teknik Lingkungan", "Fakultas Teknik - Teknik Perkapalan", 
    "Fakultas Teknik - Teknik Geologi", "Fakultas Teknik - Teknik Geodesi", 
    "Fakultas Teknik - Teknik Komputer", "Fakultas Teknik - Perencanaan Wilayah dan Kota",
    "Fakultas Ekonomika dan Bisnis - Manajemen", "Fakultas Ekonomika dan Bisnis - Ekonomi Pembangunan",
    "Fakultas Ekonomika dan Bisnis - Ekonomi Islam", "Fakultas Ekonomika dan Bisnis - Akuntansi",
    "Fakultas Hukum - Ilmu Hukum",
    "Fakultas Ilmu Budaya - Sastra Indonesia", "Fakultas Ilmu Budaya - Sastra Inggris", 
    "Fakultas Ilmu Budaya - Bahasa dan Kebudayaan Jepang", "Fakultas Ilmu Budaya - Sejarah", 
    "Fakultas Ilmu Budaya - Ilmu Perpustakaan", "Fakultas Ilmu Budaya - Antropologi Sosial",
    "Fakultas Ilmu Sosial dan Ilmu Politik - Administrasi Bisnis", 
    "Fakultas Ilmu Sosial dan Ilmu Politik - Administrasi Publik", 
    "Fakultas Ilmu Sosial dan Ilmu Politik - Ilmu Komunikasi",
    "Fakultas Ilmu Sosial dan Ilmu Politik - Ilmu Pemerintahan", 
    "Fakultas Ilmu Sosial dan Ilmu Politik - Hubungan Internasional",
    "Fakultas Kedokteran - Kedokteran", "Fakultas Kedokteran - Keperawatan", 
    "Fakultas Kedokteran - Gizi", "Fakultas Kedokteran - Kedokteran Gigi", 
    "Fakultas Kedokteran - Farmasi",
    "Fakultas Kesehatan Masyarakat - Kesehatan Masyarakat", 
    "Fakultas Kesehatan Masyarakat - Keselamatan dan Kesehatan Kerja",
    "Fakultas Peternakan dan Pertanian - Peternakan", 
    "Fakultas Peternakan dan Pertanian - Teknologi Pangan", 
    "Fakultas Peternakan dan Pertanian - Agroekoteknologi", 
    "Fakultas Peternakan dan Pertanian - Agribisnis",
    "Fakultas Perikanan dan Ilmu Kelautan - Akuakultur", 
    "Fakultas Perikanan dan Ilmu Kelautan - Ilmu Kelautan", 
    "Fakultas Perikanan dan Ilmu Kelautan - Manajemen Sumber Daya Perairan", 
    "Fakultas Perikanan dan Ilmu Kelautan - Oseanografi", 
    "Fakultas Perikanan dan Ilmu Kelautan - Perikanan Tangkap", 
    "Fakultas Perikanan dan Ilmu Kelautan - Teknologi Hasil Perikanan",
    "Fakultas Psikologi - Psikologi",
    "Fakultas Sains dan Matematika - Matematika", "Fakultas Sains dan Matematika - Biologi",
    "Fakultas Sains dan Matematika - Kimia", "Fakultas Sains dan Matematika - Fisika",
    "Fakultas Sains dan Matematika - Statistika", "Fakultas Sains dan Matematika - Ilmu Komputer/Informatika",
    "Fakultas Sains dan Matematika - Bioteknologi"
]


# Membuat comboboxes secara dinamis menggunakan perulangan
combobox_positions = [(600, 150), (600, 200)]  # Daftar posisi combobox
comboboxes = []  # List untuk menyimpan objek combobox

for i, (x, y) in enumerate(combobox_positions):
    combobox = Combobox(obj, values=jurusan_undip, font="Roboto 10", width=25, state="readonly")
    combobox.place(x=x, y=y)
    combobox.set(f"Pilihan {i + 1}")
    comboboxes.append(combobox)  # Menyimpan combobox ke dalam daftar

# Memastikan comboboxes terisi sebelum akses
if len(comboboxes) >= 2:
    Jurusan1 = comboboxes[0]
    Jurusan2 = comboboxes[1]
else:
    print("Error: Comboboxes belum terisi dengan cukup elemen.")

# Data Orang Tua
obj2=LabelFrame(root,text="Data Orang Tua",font=20,bd=2,width=1000,bg=framebg,fg=framefg,height=200,relief=GROOVE)
obj2.place(x=30,y=570)

Label(obj2,text="Nama Ayah:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Pekerjaan:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

F_Name=StringVar()
f_entry = Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)

Father_Occupation=StringVar()
FO_entry = Entry(obj2,textvariable=Father_Occupation,width=20,font="arial 10")
FO_entry.place(x=160,y=100)

Label(obj2,text="Nama Ibu:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Pekerjaan:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

M_Name=StringVar()
M_entry = Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
M_entry.place(x=630,y=50)

Mother_Occupation=StringVar()
MO_entry = Entry(obj2,textvariable=Mother_Occupation,width=20,font="arial 10")
MO_entry.place(x=630,y=100)

# Frame untuk menampilkan gambar
f = Frame(root, bd=3, bg="#0078C8", width=200, height=200, relief=GROOVE)
f.place(x=1100, y=150)

# Load dan ubah ukuran gambar
img_path = "images/upload photo.png"
image = Image.open(img_path)
image = image.resize((180, 180), Image.LANCZOS)
img = ImageTk.PhotoImage(image)

# Menampilkan gambar di Label
lbl = Label(f, bg="#0078C8", image=img)
lbl.place(x=1, y=0)

# Button
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue",command=showimage).place(x=1100, y=370)

saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen",command=Save)
saveButton.place(x=1100, y=450)

Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink",command=Clear).place(x=1100, y=530)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1100, y=690)

Button(root, text="Undo", width=19, height=2, font="arial 12 bold", bg="orange", command=undo).place(x=1100, y=610 )

root.mainloop()
